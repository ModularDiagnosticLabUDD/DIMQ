from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

mv_ = 600
t_  = 30

class calRutine():
    def __init__(self):
        self.m = 0
        self.n = 0
        self.ph = 0

        self.sensor_type = [{"ID":0,"filename": "sensor_ph_1" },
                            {"ID":1,"filename": "sensor_ph_2" },
                            {"ID":2,"filename": "sensor_ph_3" },
                            {"ID":3,"filename": "sensor_ph_4" },
                            {"ID":4,"filename": "sensor_ph_1" },
                            {"ID":6,"filename": "sensor_ph_1" },
                            {"ID":5,"filename": "sensor_ph_1" }]

        self.phs    = []    # array of phs
        self.mvs    = []    # array of millivolts
        self.mvst   = []    # matrix of millivotls
        self.temps  = []    # array of temps
        self.mv0_   = 0
        self.mv1_   = 0
        self.ph0_   = 0     
        self.ph1_   = 0 
        self.mvs0_  = []
        self.mvs1_  = []


    
    def uploadSimpleTable(self, type=0):
        for d  in self.sensor_type:
            if d['ID'] == type:
                filename_ = 'cal/'+d['filename']+'.xlsx'
                #break
        try:
            wb_ = load_workbook(filename = filename_)
            ws_ = wb_['simple']
            ph_row_ = ws_[1]
            mv_row_ = ws_[2]
            self.phs = []
            self.mvs = []
            for r in range(len(ph_row_)):
                self.phs.append( float(ph_row_[r].value))
                self.mvs.append( float(mv_row_[r].value))
            return self.phs,self.mvs
        except Exception as e:
            print("can't open file")
    
    def uploadTempTable(self, type=0):
        for d  in self.sensor_type:
            if d['ID'] == type:
                filename_ = 'cal/'+d['filename']+'.xlsx'
                #break
        try:
            wb_ = load_workbook(filename = filename_)
            ws_ = wb_['temp']
            ph_row_  = ws_[1]
            self.temps   = []
            self.mvst    = []
            self.phs     = []

            for c in ph_row_:
                if c.value is not None :
                    self.phs.append(float(c.value))
    
            for r in range(2,ws_.max_row+1):
                mv_row_= ws_[r]
                self.temps.append(float(mv_row_[0].value))
                mvs_aux_ = []
                for c in range(1,len(mv_row_)):
                    mvs_aux_.append(float(mv_row_[c].value))
                self.mvst.append(mvs_aux_)
            return self.phs,self.temps,self.mvst
        except Exception as e:
            print("can't open file")

    def pHfromData(self, pha , mva , mv_value):
        mvs_len = len(mva)
        self.mv1_ = mva[1]
        for r in range( 1 , mvs_len):
            if(mv_value < self.mv1_) or ( (r+1) > mvs_len):
                self.mv0_ = mva[r-1]
                self.ph0_ = pha[r-1]
                self.ph1_ = pha[r]
                self.m = (self.ph1_-self.ph0_)/(self.mv1_-self.mv0_)
                self.n = self.ph0_
                self.ph = self.m * (mv_value - self.mv0_) + self.ph0_
                return self.m,self.n,self.ph
            else:
                self.mv1_ = mva[r+1]
    
    def AdjustMVwithTemp(self, tempa, mvm ,temp):
        self.mvs0_ = []
        self.mvs1_ = []
        self.t0_ = 0
        self.t1_ = 0
        self.mvs = []

        if temp < tempa[0]:
            self.mvs = mvm[0]
        elif temp >= tempa[-1]:
            self.mvs = mvm[-1]  
        else:
            for idx in range(1,len(tempa)):
                if temp < tempa[idx]:
                    self.t0_ = tempa[idx-1]
                    self.t1_ = tempa[idx]
                    self.mvs0_ = mvm[idx-1]
                    self.mvs1_ = mvm[idx]
                    break
            for i in range(len(self.mvs1_)):
                self.mvs.append( ((self.mvs1_[i] - self.mvs0_[i])/(self.t1_ - self.t0_) )*(temp - self.t0_) + self.mvs0_[i] ) 
        return self.mvs

    def calculatepH(self,  mv,type= 0):
        self.uploadSimpleTable(type)
        self.pHfromData(pha = self.phs , mva = self.mvs , mv_value=mv)
        return self.ph,self.m,self.n

    def calculatephWhitTemp(self, mv, temp,type = 0 ):
        self.uploadTempTable(type)
        self.AdjustMVwithTemp(tempa = self.temps, mvm=self.mvst ,temp=temp)
        self.pHfromData(pha = self.phs , mva = self.mvs , mv_value=mv)
        return self.ph,self.m,self.n
"""
wb = load_workbook(filename = 'cal/sensor_ph_1.xlsx')
ws = wb['simple']
mv1 = float(ws.cell(row = 2, column = 2).value) #al last we need two points
for r in range(2,ws.max_row+1):
    print(f'iteración {r}')
    if (mv <= mv1) or ( (r + 1) == ws.max_row ):
        mv0 = float(ws.cell(row = r-1, column = 2).value)
        ph0 = float(ws.cell(row = r-1, column = 1).value)
        ph1 = float(ws.cell(row = r, column = 1).value)
        m   = (ph1-ph0)/(mv1-mv0)
        n   = ph0
        ph  = m * (mv - mv0) + ph0
        break
    else:
        mv1 = float(ws.cell(row = r+1, column = 2).value) #al last we need two points
"""


if __name__ == "__main__":
    mv_ = 161
    t_  = 25

    calBot = calRutine()
    print(calBot.calculatepH(type=0,mv=mv_))
    print(calBot.calculatephWhitTemp(type=0,mv=mv_,temp=t_))
    

"""


# Cargamos la tabla de temperaturas
wst     = wb['temp']
ph_row  = wst[1]
temps   = []
mvst    = []
mvs     = []
phs     = []

for c in ph_row:
    if c.value is not None :
        phs.append(float(c.value))
    


for r in range(2,wst.max_row+1):
    mv_row = wst[r]
    temps.append((mv_row[0].value))
    mvs_aux = []
    for c in range(1,len(mv_row)):
        mvs_aux.append(float(mv_row[c].value))
    mvst.append(mvs_aux)

print(temps)
print(phs)
print(mvst)
mvs0 = []
mvs1 = []
t0 = 0
t1 = 0

if t_ < temps[0]:
    mvs = mvst[0]
elif t_ >temps[-1]:
    mvs = mvst[-1]  
else:
    for idx in range(1,len(temps)):
        if t_ < temps[idx]:
            t0 = temps[idx-1]
            t1 = temps[idx]
            mvs0 = mvst[idx-1]
            mvs1 = mvst[idx]
            break
    print(mvs0)
    print(mvs1)
    for i in range(len(mvs1)):
        mvs.append( ((mvs1[i] - mvs0[i])/(t1 - t0) )*(t_ - t0) + mvs0[i] )

print(mvs)
"""